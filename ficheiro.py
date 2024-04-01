from openpyxl import load_workbook, Workbook
from crypto import value_decrypto

filename = 'Arquivo.xlsx'

try:
    ficheiro = load_workbook(filename=filename)
    sheet = ficheiro.active
except FileNotFoundError:
    # Se o arquivo não existir, cria-se um novo
    workbook = Workbook()
    sheet = workbook.active
    workbook.save(filename)
    ficheiro = load_workbook(filename=filename)
    sheet = ficheiro.active


def nova_conta(novo_mail, nova_pass, nome_util):
    print(f'Criado novo utilizador {nome_util}')
    proxima_linha = sheet.max_row + 1
    sheet.cell(row=proxima_linha, column=1,
               value=novo_mail)
    sheet.cell(row=proxima_linha, column=2,
               value=nova_pass)
    sheet.cell(row=proxima_linha, column=3,
               value=nome_util)

    ficheiro.save(filename='Arquivo.xlsx')


def verifica_conta(email, password):
    encontrado = "False"
    nome = ''
    for row in sheet.iter_rows(values_only=True):
        util = row[0]
        passwrd = row[1]
        nome_util = row[2]
        if util is not None:
            if value_decrypto(util) == value_decrypto(email) and value_decrypto(passwrd) == value_decrypto(password):
                encontrado = "True"
                nome = nome_util

    print(f'Cliente encontrado? - {encontrado} - {nome}')
    return encontrado, nome
